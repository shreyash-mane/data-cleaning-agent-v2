"""
Run cleaning pipeline and export both RAW and CLEANED data to Excel.
- Sheet 1 (Raw Data)     : original file, problem cells highlighted in red
- Sheet 2 (Cleaned Data) : cleaned output, changed cells highlighted in green
- Sheet 3 (Change Log)   : every change made, column by column
- Sheet 4 (Summary)      : model decisions and stats
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from cleaning_agent.pipeline import CleaningPipeline

RAW_PATH    = "data/raw_excel_test.csv"
OUTPUT_PATH = "data/cleaning_report.xlsx"

# ── Run pipeline ──────────────────────────────────────────────────────────────
raw      = pd.read_csv(RAW_PATH)
pipeline = CleaningPipeline()
result   = pipeline.run(raw)
cleaned  = result["cleaned_df"]
report   = result["report"]
summary  = result["summary"]

# Core columns only (exclude review_required / review_notes from main sheets)
core_cols = list(raw.columns)

# ── Colour palette ────────────────────────────────────────────────────────────
RED_BG     = PatternFill("solid", fgColor="FF4C4C")   # problem cell in raw
ORANGE_BG  = PatternFill("solid", fgColor="FFB347")   # missing in raw
GREEN_BG   = PatternFill("solid", fgColor="90EE90")   # fixed/filled in cleaned
BLUE_BG    = PatternFill("solid", fgColor="87CEEB")   # flagged in cleaned
HEADER_BG  = PatternFill("solid", fgColor="2E4057")
SUBHDR_BG  = PatternFill("solid", fgColor="4A7C9E")
ALT_BG     = PatternFill("solid", fgColor="F5F9FF")
WHITE_BG   = PatternFill("solid", fgColor="FFFFFF")

HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
BOLD_FONT    = Font(bold=True, size=10)
NORMAL_FONT  = Font(size=10)
FLAG_FONT    = Font(bold=True, color="CC0000", size=10)

THIN = Side(style="thin", color="CCCCCC")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

def style_header_row(ws, row, cols):
    for col_idx in range(1, cols + 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.fill   = HEADER_BG
        cell.font   = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER

def auto_width(ws, min_w=10, max_w=35):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_w), max_w)

# Detect which cells changed between raw and cleaned
def cell_changed(col, row_idx):
    rv = raw[col].iloc[row_idx]
    cv = cleaned[col].iloc[row_idx] if col in cleaned.columns else rv
    rs = "" if (isinstance(rv, float) and pd.isna(rv)) else str(rv).strip()
    cs = "" if (isinstance(cv, float) and pd.isna(cv)) else str(cv).strip()
    return rs, cs, rs != cs

def get_change_type(rs, cs):
    if rs == "" and cs != "":  return "FILLED"
    if rs != "" and cs == "":  return "FLAGGED"
    if rs != cs:               return "FIXED"
    return "OK"

# Detect known-bad raw values
def is_bad_raw(col, row_idx):
    rv = raw[col].iloc[row_idx]
    if isinstance(rv, float) and pd.isna(rv): return "missing"
    s = str(rv).strip().lower()
    if s in ("n/a", "na", "abc", "null", "none", "invalid-date", "not-a-date", "invalid"):
        return "invalid"
    return ""

# ─────────────────────────────────────────────────────────────────────────────
wb = Workbook()

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — RAW DATA
# ═══════════════════════════════════════════════════════════════════════════════
ws_raw = wb.active
ws_raw.title = "Raw Data"
ws_raw.freeze_panes = "A2"

# Header
for ci, col in enumerate(core_cols, 1):
    ws_raw.cell(row=1, column=ci, value=col)
style_header_row(ws_raw, 1, len(core_cols))

# Data rows
for ri, (_, row) in enumerate(raw.iterrows(), 2):
    bg = ALT_BG if ri % 2 == 0 else WHITE_BG
    for ci, col in enumerate(core_cols, 1):
        val = row[col]
        cell = ws_raw.cell(row=ri, column=ci)
        cell.value  = None if (isinstance(val, float) and pd.isna(val)) else val
        cell.font   = NORMAL_FONT
        cell.border = THIN_BORDER
        cell.alignment = LEFT

        bad = is_bad_raw(col, ri - 2)
        if bad == "missing":
            cell.fill = ORANGE_BG
            cell.value = "(missing)"
            cell.font  = Font(italic=True, color="885500", size=10)
        elif bad == "invalid":
            cell.fill = RED_BG
            cell.font = Font(bold=True, color="FFFFFF", size=10)
        else:
            cell.fill = bg

ws_raw.row_dimensions[1].height = 22
auto_width(ws_raw)

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — CLEANED DATA
# ═══════════════════════════════════════════════════════════════════════════════
ws_clean = wb.create_sheet("Cleaned Data")
ws_clean.freeze_panes = "A2"

# Header
for ci, col in enumerate(core_cols, 1):
    ws_clean.cell(row=1, column=ci, value=col)
style_header_row(ws_clean, 1, len(core_cols))

# Data rows
for ri in range(len(raw)):
    bg = ALT_BG if (ri + 2) % 2 == 0 else WHITE_BG
    for ci, col in enumerate(core_cols, 1):
        if col not in cleaned.columns:
            continue
        cv = cleaned[col].iloc[ri]
        rs, cs, changed = cell_changed(col, ri)
        change_type = get_change_type(rs, cs)

        cell = ws_clean.cell(row=ri + 2, column=ci)
        cell.value  = None if (isinstance(cv, float) and pd.isna(cv)) else cv
        cell.font   = NORMAL_FONT
        cell.border = THIN_BORDER
        cell.alignment = LEFT

        if change_type == "FILLED":
            cell.fill = GREEN_BG
            cell.font = Font(bold=True, color="006400", size=10)
        elif change_type == "FIXED":
            cell.fill = GREEN_BG
            cell.font = Font(bold=True, color="006400", size=10)
        elif change_type == "FLAGGED":
            cell.fill = BLUE_BG
            cell.value = "(flagged)"
            cell.font  = FLAG_FONT
        else:
            cell.fill = bg

ws_clean.row_dimensions[1].height = 22
auto_width(ws_clean)

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — CHANGE LOG
# ═══════════════════════════════════════════════════════════════════════════════
ws_log = wb.create_sheet("Change Log")
ws_log.freeze_panes = "A2"

log_headers = ["Row", "Employee ID", "Name", "Column", "Raw Value", "Cleaned Value", "Change Type"]
for ci, h in enumerate(log_headers, 1):
    ws_log.cell(row=1, column=ci, value=h)
style_header_row(ws_log, 1, len(log_headers))

log_row = 2
type_fills = {
    "FILLED":  PatternFill("solid", fgColor="DFFFDF"),
    "FIXED":   PatternFill("solid", fgColor="D0F0FF"),
    "FLAGGED": PatternFill("solid", fgColor="FFE0E0"),
}
type_fonts = {
    "FILLED":  Font(color="006400", bold=True, size=10),
    "FIXED":   Font(color="003580", bold=True, size=10),
    "FLAGGED": Font(color="CC0000", bold=True, size=10),
}

for ri in range(len(raw)):
    emp_id = raw["emp_id"].iloc[ri]
    name   = raw["full_name"].iloc[ri] if pd.notna(raw["full_name"].iloc[ri]) else "(unknown)"
    for col in core_cols:
        if col not in cleaned.columns:
            continue
        rs, cs, changed = cell_changed(col, ri)
        if not changed:
            continue
        change_type = get_change_type(rs, cs)
        fill = type_fills.get(change_type, WHITE_BG)
        font = type_fonts.get(change_type, NORMAL_FONT)

        vals = [ri + 1, emp_id, name, col,
                rs if rs else "(missing)", cs if cs else "(null/flagged)",
                change_type]
        for ci, v in enumerate(vals, 1):
            cell = ws_log.cell(row=log_row, column=ci, value=v)
            cell.fill   = fill
            cell.font   = font if ci == 7 else NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = LEFT
        log_row += 1

auto_width(ws_log)

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — SUMMARY
# ═══════════════════════════════════════════════════════════════════════════════
ws_sum = wb.create_sheet("Summary")

def write_section_header(ws, row, text, cols=5):
    cell = ws.cell(row=row, column=1, value=text)
    cell.fill      = SUBHDR_BG
    cell.font      = Font(bold=True, color="FFFFFF", size=11)
    cell.alignment = LEFT
    cell.border    = THIN_BORDER
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    return row + 1

def write_kv(ws, row, key, val, bold_val=False):
    kc = ws.cell(row=row, column=1, value=key)
    vc = ws.cell(row=row, column=2, value=val)
    kc.font = Font(bold=True, size=10); kc.border = THIN_BORDER; kc.alignment = LEFT
    vc.font = Font(bold=bold_val, size=10); vc.border = THIN_BORDER; vc.alignment = LEFT
    kc.fill = ALT_BG; vc.fill = ALT_BG
    return row + 1

r = 1
r = write_section_header(ws_sum, r, "  PIPELINE SUMMARY")
r = write_kv(ws_sum, r, "Total Rows",              summary["rows_before"])
r = write_kv(ws_sum, r, "Columns Processed",       summary["columns_processed"])
r = write_kv(ws_sum, r, "Columns Skipped (ID)",    summary["columns_skipped"])
r = write_kv(ws_sum, r, "Total Cells Changed",     summary["total_cells_changed"], True)
r = write_kv(ws_sum, r, "Total Cells Flagged",     summary["total_cells_flagged"])
r = write_kv(ws_sum, r, "Duplicate Rows Dropped",  summary["duplicate_rows_dropped"])
r = write_kv(ws_sum, r, "ML Decisions",            f"{summary['ml_used_count']} / {summary['columns_processed'] - summary['columns_skipped']}")
r += 1

r = write_section_header(ws_sum, r, "  PER-COLUMN MODEL DECISIONS")
col_hdr = ["Column", "Detected Type", "Action Taken", "Confidence", "Cells Changed", "Cells Flagged", "Intelligence Note"]
for ci, h in enumerate(col_hdr, 1):
    cell = ws_sum.cell(row=r, column=ci, value=h)
    cell.fill = HEADER_BG; cell.font = HEADER_FONT
    cell.alignment = CENTER; cell.border = THIN_BORDER
r += 1

action_colors = {
    "impute_mean":          "FFF8DC",
    "impute_median":        "FFF8DC",
    "convert_and_flag":     "FFE4B5",
    "parse_and_flag":       "E6F3FF",
    "normalize_and_flag":   "FFE8E8",
    "normalize_boolean":    "E8F5E9",
    "flag_outliers":        "FFE8E8",
    "fill_mode":            "F0FFF0",
    "fill_unknown":         "F0FFF0",
    "no_change":            "F5F5F5",
    "skipped":              "EEEEEE",
}

for col_report in report:
    action  = col_report["action"]
    ml_act  = col_report.get("ml_action", action)
    refined = col_report.get("intelligence_refined", False)
    note    = col_report.get("intelligence_reason", "") if refined else ""
    conf    = col_report.get("confidence", 0)
    row_color = action_colors.get(action, "FFFFFF")
    fill    = PatternFill("solid", fgColor=row_color)
    vals    = [
        col_report["column"], col_report["col_type"], action,
        f"{conf:.0%}", col_report.get("cells_changed", 0),
        col_report.get("cells_flagged", 0), note
    ]
    for ci, v in enumerate(vals, 1):
        cell = ws_sum.cell(row=r, column=ci, value=v)
        cell.fill = fill; cell.font = NORMAL_FONT
        cell.border = THIN_BORDER; cell.alignment = LEFT
    r += 1

r += 1
r = write_section_header(ws_sum, r, "  MISSING VALUES  —  BEFORE vs AFTER")
for col in core_cols:
    before = int(raw[col].isna().sum())
    after  = int(cleaned[col].isna().sum()) if col in cleaned.columns else 0
    result_str = "All filled" if before > 0 and after == 0 else \
                 f"{after} remain (unparseable)" if after > 0 else "No missing"
    row_color = "DFFFDF" if after == 0 and before > 0 else \
                "FFE0E0" if after > 0 else "FFFFFF"
    fill = PatternFill("solid", fgColor=row_color)
    for ci, v in enumerate([col, before, after, result_str], 1):
        cell = ws_sum.cell(row=r, column=ci, value=v)
        cell.fill = fill; cell.font = NORMAL_FONT
        cell.border = THIN_BORDER; cell.alignment = LEFT
    r += 1

r += 1
r = write_section_header(ws_sum, r, "  FLAGGED ROWS  (require human review)")
if "review_required" in cleaned.columns:
    flagged = cleaned[
        cleaned["review_notes"].notna() &
        cleaned["review_notes"].astype(str).str.strip().ne("")
    ]
    if len(flagged):
        for ci, h in enumerate(["Emp ID", "Name", "Review Notes"], 1):
            cell = ws_sum.cell(row=r, column=ci, value=h)
            cell.fill = HEADER_BG; cell.font = HEADER_FONT
            cell.alignment = CENTER; cell.border = THIN_BORDER
        r += 1
        for _, row in flagged.iterrows():
            fill = PatternFill("solid", fgColor="FFE0E0")
            for ci, v in enumerate([row["emp_id"], row["full_name"], row["review_notes"]], 1):
                cell = ws_sum.cell(row=r, column=ci, value=v)
                cell.fill = fill; cell.font = NORMAL_FONT
                cell.border = THIN_BORDER; cell.alignment = LEFT
            r += 1

ws_sum.column_dimensions["A"].width = 22
ws_sum.column_dimensions["B"].width = 16
ws_sum.column_dimensions["C"].width = 22
ws_sum.column_dimensions["D"].width = 12
ws_sum.column_dimensions["E"].width = 14
ws_sum.column_dimensions["F"].width = 14
ws_sum.column_dimensions["G"].width = 40

# ── Save ──────────────────────────────────────────────────────────────────────
wb.save(OUTPUT_PATH)

# ── Console summary ───────────────────────────────────────────────────────────
print(f"\n{'='*60}")
print(f"  Excel report saved -> {OUTPUT_PATH}")
print(f"{'='*60}")
print(f"  Sheets:")
print(f"    Raw Data     — original 30 rows, problem cells in RED/ORANGE")
print(f"    Cleaned Data — fixed values in GREEN, flagged in BLUE")
print(f"    Change Log   — every change row by row")
print(f"    Summary      — model decisions + missing value scorecard")
print(f"\n  Total cells changed : {summary['total_cells_changed']}")
print(f"  Total cells flagged : {summary['total_cells_flagged']}")
print(f"  ML decisions        : {summary['ml_used_count']}/{summary['columns_processed']-summary['columns_skipped']}")

print(f"\n  Per-column decisions:")
for r in report:
    ml  = r.get("ml_action", r["action"])
    fin = r["action"]
    ref = "  [refined]" if r.get("intelligence_refined") else ""
    chg = r.get("cells_changed", 0)
    flg = r.get("cells_flagged", 0)
    print(f"    {r['column']:<18} {r['col_type']:<10} {fin:<22} chg={chg}  flg={flg}{ref}")

print()
